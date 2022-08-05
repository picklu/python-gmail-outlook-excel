import os
import openpyxl

from datetime import datetime
from decouple import config
from gmail import GMail, Message


wb_folder = config('WORKING_FOLDER')
dt_folder = config('DATA_FOLDER')
wb_name = config('WB_NAME')
wb_path = os.path.join(wb_folder, wb_name)
workbook = openpyxl.load_workbook(wb_path)


class Student:

    def __init__(self, department):
        self.department = department
        self.id = None
        self.name = None
        self.email = None
        self.mobile = None
        self.paid = None
        self.file_path = None

    def __repr__(self) -> str:
        return f"__{self.name}({self.id}) of {self.department}"


def send_gmail(mail, student):
    """Send email using gmail. Use your App Password 
    to avoid any error sending Gmail
    """

    mail_property = {}
    mail_property["subject"] = f"Final result for Spring 2022"
    mail_property["to"] = student.email
    mail_property["cc"] = config('CC_MAIL_ID')
    mail_property["text"] = "Final result for Spring 2020 is availabe now"
    mail_property["html"] = f"""
    Dear {student.name},<br><br>
    The final results for Spring 2022 is available now.<br>
    Please find attahced herewith your result.
    <br><br>
    All the best.
    <br><br>
    Dr. Subrata Sarker<br>
    Registrar<br>
    University of Skill Enrichment and Technology<br>
    Signboard Circle, Narayanganj, Dhaka, Bangladesh<br>
    e-mail: ss.rgstr.uset.edu@gmail.com<br>
    Dispatched at {datetime.now().strftime("%d %B %Y, %H:%M:%S %p")}
    """
    mail_property["attachments"] = [student.file_path]
    msg = Message(**mail_property)
    mail.send(msg)

    print(f"==> Mail sent to {student.name}<{student.email}> [{student.id}]")


if __name__ == "__main__":

    mail = GMail(
        f"Dr. Subrata Sarker<{config('EMAIL_ID')}>", config('PASSWORD'))

    ws_names = ["CSE", "English"]
    for ws_name in ws_names:
        ws = workbook[ws_name]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == None:
                continue

            student = Student(ws_name)

            for cell in row:
                match ws.cell(1, cell.column).value.lower():
                    case "student id":
                        student.id = cell.value
                    case "name":
                        student.name = cell.value
                    case "email":
                        student.email = cell.value
                    case "mobile":
                        student.mobile = cell.value
                    case "paid":
                        student.paid = cell.value == "Paid"
                    case "file name":
                        student.file_path = os.path.join(
                            dt_folder, cell.value)

            if student.name and student.paid:
                print(
                    f"==> Mail to be sent to {student.name}<{student.email}> [{student.id}]")
                send_gmail(mail, student)
            else:
                print(
                    f"==O Mail not to be sent to {student.name}<{student.email}> [{student.id}]")
    print("==> Done!")
